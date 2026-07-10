from playwright.sync_api import Playwright,Page,expect
from openpyxl import load_workbook
import pytest

# workbook=load_workbook("test_data_DDT\Book (1).xlsx")
# sheet=workbook["Sheet1"]

# TEST_DATA=[]

# for row in sheet.iter_rows(values_only=True):
#     TEST_DATA.append(row)
# print(TEST_DATA)

# @pytest.mark.parametrize("Name,Age",TEST_DATA)
# def test_data(Name,Age):
#     print(Name,Age)


# workbook=load_workbook("test_data_DDT\Marks.xlsx.xlsx")
# sheet=workbook["Sheet1"]
# test_data_002=[]

# for row in sheet.iter_rows(min_row=2,values_only=True):
#     test_data_002.append(row)
# print(test_data_002)
# @pytest.mark.parametrize("Student,Marks",test_data_002)
# def test_get_result(Student,Marks):
#     if Marks==100:
#         result="Excellent Performance"
#     elif Marks>=50:
#         result="Passed"
#     else:
#         result="Failed"
#     print(f"{Student} is {result}")


# workbook=load_workbook("test_data_DDT\LoginData.xlsx")
# sheet=workbook["Sheet1"]
# test_data_003=[]

# for row in sheet.iter_rows(min_row=2,values_only=True):
#     test_data_003.append(row)
# print(test_data_003)
# def login(Username, Password):
#     if Username=="admin" and Password=="admin123":
#         return "Success"
#     elif Username !="admin":
#         return "Failed"
#     else:
#         return "Failed"
    
# @pytest.mark.parametrize("Username,Password,Expected",test_data_003)
# def test_login_using_excel(Username,Password,Expected):
#     assert login(Username, Password)== Expected


workbook=load_workbook("test_data_DDT\PLAYWRIGHT.xlsx")
sheet=workbook["Sheet1"]
test_data_004=[]

for row in sheet.iter_rows(min_row=2,values_only=True):
    test_data_004.append(row)
print(test_data_004)


@pytest.mark.parametrize("Username,Password",test_data_004)

def test_assignment_004(page:Page,Username,Password):

#     # Open login page
    page.goto("https://practicetestautomation.com/practice-test-login/")

#     # Enter Username
    page.locator("input[id='username'][name='username']").fill(Username)

#     # Enter Password
    page.locator("input[id='password'][name='password']").fill(str(Password))

#     # Click Login button
    page.locator("button[id='submit']").click()

#     # Scenario 1 : Valid Login
    if Username=="student" and Password == "Password123":

#         # Print current URL
        print(page.url)

#         # Verify successful navigation
        expect(page).to_have_url("https://practicetestautomation.com/logged-in-successfully/")

#         # Verify success heading
        expect(page.locator(".post-title")).to_contain_text("Logged In Successfully")

#         # Verify Logout link is visible
        expect(page.get_by_role("link", name="Log out")).to_be_visible()

#         # Capture success message
        success=page.locator(".post-title").text_content()
        print(f"Application result:{success}")

#     # Scenario 2 : Invalid Username
    elif Username !="student":

#         # Verify username error message
        expect(page.locator("div[id='error']")).to_have_text("Your username is invalid!")

        # Capture error message
        error = page.locator("div[id='error']").text_content()
        print(f"Application Result: {error}")

#     # Scenario 3 : Invalid Password
    else:

#         # Verify password error message
        expect(page.locator("div[id='error']")).to_have_text("Your password is invalid!")

#         # Capture error message
        error = page.locator("div[id='error']").text_content()
        print(f"Application Result: {error}")


